#!/bin/env python
# -*- coding:utf-8 -*-

r"""
CSV 文件转 XLSX 文件
- CSV 文件编码必须 UTF-8
- Windows系统 文件路径 需用 \\ 替代 \ 以免转义, 或者用单引号 ' 包起来
"""

import csv

import click
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def read_csv(file_name):
    with open(file_name) as f:
        rd = csv.reader(f)
        content = [r for r in rd]
    return content


def push2xlsx(sheet, file_name):
    content = read_csv(file_name)
    for r, row in enumerate(content, 2):
        for c, cell in enumerate(row, 1):
            sheet["%s%s" % (get_column_letter(c), r)] = cell


def csv2xlsx(workbook, sheet, csv_file, xlsx_file):
    push2xlsx(sheet, csv_file)
    workbook.save(xlsx_file)


@click.command()
@click.option("--source", required=True, help=u"csv 文件名")
@click.option("--target", default="", help=u"xlsx 文件名, 默认使用与 source 同名但以 .xlsx 后缀的文件")
@click.option("--sheet", default="", help=u"表名")
@click.option("--append", is_flag=True, help=u"在 target 文件的新 sheet 上追加 source 内容")
def main(source, target, sheet, append):
    if not target:
        target = source.split(".")[0] + ".xlsx"
    if append:
        wb = load_workbook(target)
        sh = wb.create_sheet(sheet or None)
        csv2xlsx(wb, sh, source, target)
    else:
        wb = Workbook()
        sh = wb.active
        if sheet:
            sh.title = sheet
        csv2xlsx(wb, sh, source, target)


if __name__ == "__main__":
    main()